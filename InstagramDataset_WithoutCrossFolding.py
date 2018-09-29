def finding_coeficients() :
    import os #for giving adress to where we want to save output files
    from openpyxl import load_workbook
    wb = load_workbook('MainDataset.xlsx')
    ws = wb.active
    name=[]
    number_of_posts = []
    number_of_followers = []
    number_of_followings = []
    number_of_likes_for_10th_ex_post = []
    number_of_likes_for_11th_ex_post = []
    number_of_likes_for_12th_ex_post = []
    number_of_self_picture_posts_form_9_previous_posts = []
    sex = []
    
    #getting data of each column of excel file into a list named by the name of corrospondant name of that column 
    for row in ws :
        if row[2].value != None :  # because after 150 days the profile of some users were not availbale so we need to delete their rows by this line (look at the dataset)
            name.append(row[0].value)
            number_of_posts.append(row[1].value)
            number_of_followers.append(row[2].value)
            number_of_followings.append(row[3].value)
            number_of_likes_for_10th_ex_post.append(row[4].value)
            number_of_likes_for_11th_ex_post.append(row[5].value)
            number_of_likes_for_12th_ex_post.append(row[6].value)
            number_of_self_picture_posts_form_9_previous_posts.append(row[7].value)
            sex.append(row[8].value)

#------------------------------------
    #finding the number of users in dataset
    number_of_users_in_dataset = len (name)    
    print ("number_of_users_in_dataset = ", number_of_users_in_dataset - 1)

#------------------------------------        
    #finding the number of male and female users in dataset
    i=1 #row[0] belongs to lables of columns so by setting i=1 we ignore that first row
    f_number = 0
    m_number = 0
    while i < number_of_users_in_dataset : #because total number of users is 1000 and the first row is for labels so we have to go up to 1001
        if sex[i] == "f" :
            f_number += 1

        elif sex[i] == "m" :
            m_number += 1
        else :
            print("aha")


        i+=1
    print ("f_number :",f_number)
    print ("m_number :",m_number)
#------------------------------------
    #finding the mean number of self posts (from 9 posts) for male and female users in dataset
    i=1 #row[0] belongs to lables of columns so by setting i=1 we ignore that first row
    sum_self_post_for_all_females = 0
    sum_self_post_for_all_males = 0
    while i < number_of_users_in_dataset : 
        if sex[i] == "f" :
            t = number_of_self_picture_posts_form_9_previous_posts[i]
            sum_self_post_for_all_females += int(t)
        elif sex[i] == "m" :
            w = number_of_self_picture_posts_form_9_previous_posts[i]
            sum_self_post_for_all_males += int(w)
            
        i+=1
        
    mean_of_self_post_for_females = sum_self_post_for_all_females/f_number
    mean_of_self_post_for_males = sum_self_post_for_all_males/m_number

#-----------------------------------
    #finding the mean number of followers for male and female users in dataset
    i=1 #row[0] belongs to lables of columns so by setting i=1 we ignore that first row
    sum_number_of_followers_for_all_females = 0
    sum_number_of_followers_for_all_males = 0
    while i < number_of_users_in_dataset : 
        if sex[i] == "f" :
            t = number_of_followers[i]
            sum_number_of_followers_for_all_females += int(t)
        elif sex[i] == "m" :
            w = number_of_followers[i]
            sum_number_of_followers_for_all_males += int(w)
            
        i+=1
        
    mean_number_of_followers_for_females = sum_number_of_followers_for_all_females/f_number
    mean_number_of_followers_for_males = sum_number_of_followers_for_all_males/m_number
    
#-------------------------------------
    #finding the mean number of followings for male and female users in dataset
    i=1 #row[0] belongs to lables of columns so by setting i=1 we ignore that first row
    sum_number_of_followings_for_all_females = 0
    sum_number_of_followings_for_all_males = 0
    while i < number_of_users_in_dataset : 
        if sex[i] == "f" :
            t = number_of_followings[i]
            sum_number_of_followings_for_all_females += int(t)
        elif sex[i] == "m" :
            w = number_of_followings[i]
            sum_number_of_followings_for_all_males += int(w)
            
        i+=1
        
    mean_number_of_followings_for_females = sum_number_of_followings_for_all_females/f_number
    mean_number_of_followings_for_males = sum_number_of_followings_for_all_males/m_number

#-------------------------------------
    #finding the mean number of posts for male and female users in dataset
    i=1 #row[0] belongs to lables of columns so by setting i=1 we ignore that first row
    sum_number_of_posts_for_all_females = 0
    sum_number_of_posts_for_all_males = 0
    while i < number_of_users_in_dataset : 
        if sex[i] == "f" :
            t = number_of_posts[i]
            sum_number_of_posts_for_all_females += int(t)
        elif sex[i] == "m" :
            w = number_of_posts[i]
            sum_number_of_posts_for_all_males += int(w)
            
        i+=1
        
    mean_number_of_posts_for_females = sum_number_of_posts_for_all_females/f_number
    mean_number_of_posts_for_males = sum_number_of_posts_for_all_males/m_number
#-------------------------------------
    #finding the relation between number of self posts(in 9 previous posts) and following/follower 
    i=1 #row[0] belongs to lables of columns so by setting i=1 we ignore that first row
    ratio_of_followings_to_followers_for_all_users = []
    while i < number_of_users_in_dataset : 
        t = number_of_followings[i]/number_of_followers[i]
        ratio_of_followings_to_followers_for_all_users.append(t)   
        i+=1
    i = 0

    del number_of_self_picture_posts_form_9_previous_posts[0] #before this command, "number_of_self_picture_posts_form_9_previous_posts" was a list by 1001 elemnts, remember the first element was the label, but we need a that list to just have 1000 elemnts, so be the same size of "ratio_of_followings_to_followers_for_all_users". by having both a 1000 elements list then we can compare and analyse them.

#---------------------------------------------------------
   #finding the relation between the mean number of likes for posts(based on ex 10th 11th 12th posts) and the ratio of following/follower
    i=1
    mean_like_for_all_users = []
    while i < number_of_users_in_dataset :
        total_like = number_of_likes_for_10th_ex_post[i] + number_of_likes_for_11th_ex_post[i] + number_of_likes_for_12th_ex_post[i]
        mean_like = total_like/3
        mean_like_for_all_users.append (mean_like)
        i+=1
##    import matplotlib.pyplot as plt
##    import numpy
##    plt.hist2d(mean_like_for_all_users,ratio_of_followings_to_followers_for_all_users,50)
##    plt.colorbar()
##    plt.show()

#-------- Hist2D : like/following & following/follower
    #finding the relation between the ratio of "mean number of likes for posts(based on ex 10th 11th 12th posts)/following" and the ratio of "following/follower"
    i=0
    mean_like_to_following_for_all_users = []
    del number_of_followings[0]
    while i < number_of_users_in_dataset -1 :
        mean_like_to_following = mean_like_for_all_users[i]/number_of_followings[i]
        mean_like_to_following_for_all_users.append (mean_like_to_following)
        i+=1
    import matplotlib.pyplot as plt
    import numpy
    print("MAX[mean_like_to_following_for_all_users] = " , max(mean_like_to_following_for_all_users)) #we need this line to set max limit for horizental axis in plt.hist2d command (two lines below), which is 32.12 (we set to 33 ) , notice once we run the program and see the 15.66 then we set it in plt.hist2d for ever
    print("MAX[ratio_of_followings_to_followers_for_all_users] = " , max(ratio_of_followings_to_followers_for_all_users)) #we need this line to set max limit for vertical axis in plt.hist2d command (two lines below), which is 15.66 (we set to 16 ) , notice once we run the program and see the 15.66 then we set it in plt.hist2d for ever
    plt.hist2d(mean_like_to_following_for_all_users,ratio_of_followings_to_followers_for_all_users,100,[[0,2],[0,16]]) #notice max horizental was 33 but since there was only a few data with more than 2 as ratio of like to following so we ignored them and set the max horizental limit to 2
##    plt.scatter(mean_like_to_following_for_all_users,ratio_of_followings_to_followers_for_all_users)
    plt.xlabel('Like/Following ratio')
    plt.ylabel('Following/Follower ratio')
##    plt.title('Like/Following & Following/Follower',fontsize=12)
    file_name = "Hist2D_LikeFollowing_to_FollowingFollower"  # notice that if you set name as following/follower , I mean if you put / symbol in the name then it won't be saved! because symboles such this "/" must not be in file names
    path_of_saving = str ("E:\\Papers\\Paper_2\\Codes\\Result_Pics") #here we make the address of the folder we want pictures to be saved, BUT YOU HAVE TO douplicate each Backslash in the URL address (as you see I did ) , if you dont it wont work and give you error of UNICODESCAPE !
    cbar = plt.colorbar() #the reason why we don't use only plt.colorbar() lonely , is that we want to put a label close it, so we used this line and next line 
    cbar.set_label ("Frequency")
    plt.savefig(os.path.join(path_of_saving, file_name))
    plt.show()
#-------- Hist2D : like/follower & following/follower
    #finding the relation between the ratio of "mean number of likes for posts(based on ex 10th 11th 12th posts)/follower" and the ratio of "following/follower"
    i=0
    mean_like_to_follower_for_all_users = []
    del number_of_followers[0]
    while i < number_of_users_in_dataset - 1 :
        mean_like_to_follower = mean_like_for_all_users[i]/number_of_followers[i]
        mean_like_to_follower_for_all_users.append (mean_like_to_follower)
        i+=1
    import matplotlib.pyplot as plt
    import numpy as np
    import matplotlib.colors as mcolors
    import matplotlib.mlab as mlab
    import os #To give address for saving output plots
##    import scipy.stats

##    from numpy._distributor_init import NUMPY_MKL 
##    from scipy.stats import norm
    print("MAX[mean_like_to_follower_for_all_users] = " , max(mean_like_to_follower_for_all_users)) #we need this line to set max limit for horizental axis in plt.hist2d command (two lines below), which is 0.92 (we set to 1 ) , notice once we run the program and see the 0.92 then we set it in plt.hist2d for ever
    print("MAX[ratio_of_followings_to_followers_for_all_users] = " , max(ratio_of_followings_to_followers_for_all_users)) #we need this line to set max limit for vertical axis in plt.hist2d command (two lines below), which is 15.66 (we set to 16 ) , notice once we run the program and see the 15.66 then we set it in plt.hist2d for ever
    plt.hist2d(mean_like_to_follower_for_all_users,ratio_of_followings_to_followers_for_all_users,100,[[0,1],[0,16]])
##    counts, _ , _ , _ = plt.hist2d(mean_like_to_follower_for_all_users,ratio_of_followings_to_followers_for_all_users,100,[[0,1],[0,16]])
##    row_total = 0
##    total = 0
##    for w in counts :
##        row_total += w
##    for y in row_total :
##        total += y
##    print(total)
    plt.xlabel('Like/Follower ratio')
    plt.ylabel('Following/Follower ratio')
##    plt.title('Like/Follower & Following/Follower',fontsize=12)
    file_name = "Hist2D_LikeFollower_to_FollowingFollower"  # notice that if you set name as following/follower , I mean if you put / symbol in the name then it won't be saved! because symboles such this "/" must not be in file names
    path_of_saving = str ("E:\\Papers\\Paper_2\\Codes\\Result_Pics") #here we make the address of the folder we want pictures to be saved, BUT YOU HAVE TO douplicate each Backslash in the URL address (as you see I did ) , if you dont it wont work and give you error of UNICODESCAPE !
    cbar = plt.colorbar() #the reason why we don't use only plt.colorbar() lonely , is that we want to put a label close it, so we used this line and next line 
    cbar.set_label ("Frequency")
    plt.savefig(os.path.join(path_of_saving, file_name))
    plt.show()
    #---- Hist2D : like/follower & following/follower --- its the above hist2d but here we just focus on the dense area  
    plt.hist2d(mean_like_to_follower_for_all_users,ratio_of_followings_to_followers_for_all_users,30,[[0,0.5],[0,10]])
##    plt.scatter(mean_like_to_follower_for_all_users,ratio_of_followings_to_followers_for_all_users)
##    plt.show()
    plt.xlabel('Like/Follower ratio')
    plt.ylabel('Following/Follower ratio')
##    plt.title('Like/Follower & Following/Follower (focused)',fontsize=12)
    file_name = "Hist2Dfocused_LikeFollower_to_FollowingFollower"  # notice that if you set name as following/follower , I mean if you put / symbol in the name then it won't be saved! because symboles such this "/" must not be in file names
    path_of_saving = str ("E:\\Papers\\Paper_2\\Codes\\Result_Pics") #here we make the address of the folder we want pictures to be saved, BUT YOU HAVE TO douplicate each Backslash in the URL address (as you see I did ) , if you dont it wont work and give you error of UNICODESCAPE !
    cbar = plt.colorbar() #the reason why we don't use only plt.colorbar() lonely , is that we want to put a label close it, so we used this line and next line 
    cbar.set_label ("Frequency")
    plt.savefig(os.path.join(path_of_saving, file_name))
    plt.show()
    #-------- Hist1D : following/follower 
    plt.hist(ratio_of_followings_to_followers_for_all_users,[0,0.5,1,1.5,2,2.5,3,3.5,4,4.5,5,5.5,6,6.5,7,7.5,8,8.5,9,9.5,10,10.5,11,11.5,12,12.5,13,13.5,14,14.5,15,15.5,16]) #we set the edges of bins for histogram from 0 to 16(which is the max in the dataset and we noticed that from previous command  

    #-----
##    pdf_x = np.linspace(np.min(ratio_of_followings_to_followers_for_all_users),np.max(ratio_of_followings_to_followers_for_all_users),100)
##    pdf_y = 1.0/np.sqrt(2*np.pi*var)*np.exp(-0.5*(pdf_x-avg)**2/var)
##    plt.plot(pdf_x,pdf_y,'k--')
##    print(avg)
##    print(var)
##    plt.colorbar()
##    (mu, sigma) = norm.fit(datos)
##    y = mlab.normpdf( bins, mu, sigma)
##    l = plt.plot(bins, y, 'r--', linewidth=2)
     
    plt.xlabel('Following/Follower ratio')
    plt.ylabel('Frequency')
##    plt.title('Following/Follower',fontsize=12)
    file_name = "Hist1D_Following_to_Follower"  # notice that if you set name as following/follower , I mean if you put / symbol in the name then it won't be saved! because symboles such this "/" must not be in file names
    path_of_saving = str ("E:\\Papers\\Paper_2\\Codes\\Result_Pics") #here we make the address of the folder we want pictures to be saved, BUT YOU HAVE TO douplicate each Backslash in the URL address (as you see I did ) , if you dont it wont work and give you error of UNICODESCAPE !
    plt.savefig(os.path.join(path_of_saving, file_name))
    plt.show()
    #-------- Hist1D : like/follower

    plt.hist(mean_like_to_follower_for_all_users,[0,0.05,0.1,0.15,0.2,0.25,0.3,0.35,0.4,0.45,0.5,0.55,0.6,0.65,0.7,0.75,0.8,0.85,0.9,0.95,1]) #we set the edges of bins for histogram from 0 to 0.92(which is the max in the dataset and we noticed that from previous commands  
    #--------
##    pdf_x = np.linspace(np.min(ratio_of_followings_to_followers_for_all_users),np.max(ratio_of_followings_to_followers_for_all_users),100)
##    pdf_y = 1.0/np.sqrt(2*np.pi*var)*np.exp(-0.5*(pdf_x-avg)**2/var)
##    plt.plot(pdf_x,pdf_y,'k--')
##    print(avg)
##    print(var)
##    plt.colorbar()
##    (mu, sigma) = norm.fit(datos)
##    y = mlab.normpdf( bins, mu, sigma)
##    l = plt.plot(bins, y, 'r--', linewidth=2)
     
    plt.xlabel('Like/Follower ratio')
    plt.ylabel('Frequency')
##    plt.title('Like/Follower',fontsize=12)
    file_name = "Hist1D_Like_to_Follower"  # notice that if you set name as following/follower , I mean if you put / symbol in the name then it won't be saved! because symboles such this "/" must not be in file names
    path_of_saving = str ("E:\\Papers\\Paper_2\\Codes\\Result_Pics") #here we make the address of the folder we want pictures to be saved, BUT YOU HAVE TO douplicate each Backslash in the URL address (as you see I did ) , if you dont it wont work and give you error of UNICODESCAPE !
    plt.savefig(os.path.join(path_of_saving, file_name))
    plt.show()
    #-------- 

    

#-----------------------------------------------------------
    #great result :
    #THOSE USERS WITH FOLLOWEING/FOLLOWER < 1 WITH STRONGER PROBABILITY HAVE BETTER CHANCE OF ACCEPTANCE . In other words they have better like/follower (like/following) , so the chance that they start the rumor are more stronger 
#-----------------------------------------------------------    
    #finding the relation between the number of self posts(in 9 previous posts) (NARCISSISM) and the ratio of "mean number of likes for posts(based on ex 10th 11th 12th posts)/follower"
    #PLEASE notice that if we use discrete values of 0,1,2,3,..,9 as number of self posts(narcissism) then hist2d shape would be discrete that is not good, so for each number we randomly choose a value between two defined limit. for example if self posts number is 2 then we choose ranmoly a number between (0.2 , 0.3) to represent its narcissism level
    i=0
    import matplotlib.pyplot as plt
    import numpy
    import random
    narcissism = []
    while i < number_of_users_in_dataset - 1 :
        if number_of_self_picture_posts_form_9_previous_posts[i] == 0 :
            t = random.uniform(0,0.1)
            narcissism.append (t)
        elif number_of_self_picture_posts_form_9_previous_posts[i] == 1 :
            t = random.uniform(0.1,0.2)
            narcissism.append (t)
        elif number_of_self_picture_posts_form_9_previous_posts[i] == 2 :
            t = random.uniform(0.2,0.3)
            narcissism.append (t)
        elif number_of_self_picture_posts_form_9_previous_posts[i] == 3 :
            t = random.uniform(0.3,0.4)
            narcissism.append (t)
        elif number_of_self_picture_posts_form_9_previous_posts[i] == 4 :
            t = random.uniform(0.4,0.5)
            narcissism.append (t)
        elif number_of_self_picture_posts_form_9_previous_posts[i] == 5 :
            t = random.uniform(0.5,0.6)
            narcissism.append (t)
        elif number_of_self_picture_posts_form_9_previous_posts[i] == 6 :
            t = random.uniform(0.6,0.7)
            narcissism.append (t)
        elif number_of_self_picture_posts_form_9_previous_posts[i] == 7 :
            t = random.uniform(0.7,0.8)
            narcissism.append (t)
        elif number_of_self_picture_posts_form_9_previous_posts[i] == 8 :
            t = random.uniform(0.8,0.9)
            narcissism.append (t)
        elif number_of_self_picture_posts_form_9_previous_posts[i] == 9 :
            t = random.uniform(0.9,1)
            narcissism.append (t)
        i+=1
    plt.hist2d(narcissism,mean_like_to_follower_for_all_users,40,[[0,1],[0,1]])
    plt.xlabel('Ratio of self-presenting posts')
    plt.ylabel('Like/Follower')
##    plt.title('narcissism & Like/Follower ratio',fontsize=12)
    file_name = "Hist2D_narcissism_to_LikeFollower"  # notice that if you set name as following/follower , I mean if you put / symbol in the name then it won't be saved! because symboles such this "/" must not be in file names
    path_of_saving = str ("E:\\Papers\\Paper_2\\Codes\\Result_Pics") #here we make the address of the folder we want pictures to be saved, BUT YOU HAVE TO douplicate each Backslash in the URL address (as you see I did ) , if you dont it wont work and give you error of UNICODESCAPE !
    cbar = plt.colorbar() #the reason why we don't use only plt.colorbar() lonely , is that we want to put a label close it, so we used this line and next line 
    cbar.set_label ("Frequency")
    plt.savefig(os.path.join(path_of_saving, file_name))
    plt.show()
#---- #finding the relation between the number of self posts(in 9 previous posts) (NARCISSISM) and the ratio of following/follower"
    plt.hist2d(narcissism,ratio_of_followings_to_followers_for_all_users,40,[[0,1],[0,16]])
    plt.xlabel('Ratio of self-presenting posts')
    plt.ylabel('Following/Follower')
##    plt.title('narcissism & Following/Follower ratio',fontsize=12)
    file_name = "Hist2D_narcissism_to_FollowingFollower"  # notice that if you set name as following/follower , I mean if you put / symbol in the name then it won't be saved! because symboles such this "/" must not be in file names
    path_of_saving = str ("E:\\Papers\\Paper_2\\Codes\\Result_Pics") #here we make the address of the folder we want pictures to be saved, BUT YOU HAVE TO douplicate each Backslash in the URL address (as you see I did ) , if you dont it wont work and give you error of UNICODESCAPE !
    cbar = plt.colorbar() #the reason why we don't use only plt.colorbar() lonely , is that we want to put a label close it, so we used this line and next line 
    cbar.set_label ("Frequency")
    plt.savefig(os.path.join(path_of_saving, file_name))
    plt.show()
##----------
    #---- #finding the relation between the number of self posts(in 9 previous posts) (NARCISSISM) and the number of posts
    i=1
    del number_of_posts[0]
    plt.hist2d(narcissism,number_of_posts,40,[[0,1],[50,2000]]) #notice the min of posts was 50 because we didnt pick users with less than 50 posts and also max was 10000 but since there was a few users with more than 2000 so we made up limit to 2000
    plt.xlabel('Ratio of self-presenting posts')
    plt.ylabel('number_of_posts')
##    plt.title('narcissism & number_of_posts ratio',fontsize=12)
    file_name = "Hist2D_narcissism_to_numberofposts"  # notice that if you set name as following/follower , I mean if you put / symbol in the name then it won't be saved! because symboles such this "/" must not be in file names
    path_of_saving = str ("E:\\Papers\\Paper_2\\Codes\\Result_Pics") #here we make the address of the folder we want pictures to be saved, BUT YOU HAVE TO douplicate each Backslash in the URL address (as you see I did ) , if you dont it wont work and give you error of UNICODESCAPE !
    cbar = plt.colorbar() #the reason why we don't use only plt.colorbar() lonely , is that we want to put a label close it, so we used this line and next line 
    cbar.set_label ("Frequency")
    plt.savefig(os.path.join(path_of_saving, file_name))
    plt.show()
    #-----
    #---- #finding the following/follower ratio and the number of posts
    i=1
    plt.hist2d(ratio_of_followings_to_followers_for_all_users,number_of_posts,40,[[0,16],[50,2000]]) #notice the min of posts was 50 because we didnt pick users with less than 50 posts and also max was 10000 but since there was a few users with more than 2000 so we made up limit to 2000
    plt.xlabel('following/follower')
    plt.ylabel('number_of_posts')
    plt.title('following/follower & number_of_posts ratio',fontsize=12)
    file_name = "Hist2D_narcissism_to_numberofposts"  # notice that if you set name as following/follower , I mean if you put / symbol in the name then it won't be saved! because symboles such this "/" must not be in file names
    path_of_saving = str ("E:\\Papers\\Paper_2\\Codes\\Result_Pics") #here we make the address of the folder we want pictures to be saved, BUT YOU HAVE TO douplicate each Backslash in the URL address (as you see I did ) , if you dont it wont work and give you error of UNICODESCAPE !
    cbar = plt.colorbar() #the reason why we don't use only plt.colorbar() lonely , is that we want to put a label close it, so we used this line and next line 
    cbar.set_label ("Frequency")
    plt.savefig(os.path.join(path_of_saving, file_name))
    plt.show()
##    plt.scatter(mean_like_to_following_for_all_users,ratio_of_followings_to_followers_for_all_users)



##    c=len(x)
##    cc=len(y)
##    print("length_x",c,"length_y",cc)
##    from matplotlib import pyplot
##    import matplotlib.pyplot as plt
##    pyplot.scatter(x,y)
##    plt.show()
    
##    #showing results
##    print ("-------------------------------------------")
##    print ("-------------------------------------------")
##    print ("female stattistics : ")
##    print ("number of female users : " , f_number)
##    print ("mean_of_self_post_for_each_female : " , mean_of_self_post_for_females)
##    print ("mean_number_of_followers_for_each_female : " , mean_number_of_followers_for_females)
##    print ("mean_number_of_followings_for_each_female : " , mean_number_of_followings_for_females)
##    print ("mean_number_of_posts_for_each_female : " , mean_number_of_posts_for_females)
##    print ("-------------------------------------------")
##    print ("-------------------------------------------")
##    print ("male stattistics : ")
##    print ("number of male users : " , m_number)
##    print ("mean_of_self_post_for_each_male : " , mean_of_self_post_for_males)
##    print ("mean_number_of_followers_for_each_male : " , mean_number_of_followers_for_males)
##    print ("mean_number_of_followings_for_each_male : " , mean_number_of_followings_for_males)
##    print ("mean_number_of_posts_for_each_male : " , mean_number_of_posts_for_males)
##    print ("-------------------------------------------")
##    print ("-------------------------------------------")
###-------------------------------------    
# finding the relation of self post with  "following/follower" , "like/follower" and "sex"
# notic we want to find best amount for cc , cx , cxx , cy , cyy , cq in the equation z = cc + cx*x + cxx*x + cy*y + cyy*y + cq*q (in which x , y , q are independent variables)
    import matplotlib.pyplot as plt
    import numpy as np
    i = 1
    sex_values = []
    while i < number_of_users_in_dataset :
        if sex[i] == "m" :
            sex_values.append(10)
        elif sex[i] == "f" :
            sex_values.append(-10)
        i+=1
        
    c = np.ones(number_of_users_in_dataset -1) #we make a list c=[1,1,1,...,1] (number_of_users_in_dataset -1 elements which is same size with x , y and q) for constant amount in the equation (ie: cc) 
    x = mean_like_to_follower_for_all_users
    y = ratio_of_followings_to_followers_for_all_users
    q = sex_values


    xx = [] #we want to get xx = x**2
    xxx = [] #we want to get xxx = x**3
    for x_element in x :
        element_power_2 = x_element**2
        xx.append(element_power_2)
        element_power_3 = x_element**3
        xxx.append(element_power_3)
        
    yy = [] #we want to get yy = y**2
    yyy = [] #we want to get yyy = y**3
    for y_element in y :
        element_power_2 = y_element**2
        yy.append(element_power_2)
        element_power_3 = y_element**3
        yyy.append(element_power_3)


        


    
    z = number_of_self_picture_posts_form_9_previous_posts
    A_3 = np.column_stack ((c,x,xx,xxx,y,yy,yyy,q)) # cubic equation
    A_2 = np.column_stack ((c,x,xx,y,yy,q)) # quadratic equation
    B = z
    result_3, _ , _ , _ = np.linalg.lstsq(A_3, B)
    result_2, _ , _ , _ = np.linalg.lstsq(A_2, B)
    cc_3 , cx_3 , cxx_3 , cxxx_3, cy_3 , cyy_3 , cyyy_3 , cq_3   = result_3
    cc_2 , cx_2 , cxx_2 , cy_2 , cyy_2 , cq_2   = result_2
    print ("-------------------------------------")
    print("cc_3 = ",cc_3 , "cx_3 = ",cx_3 , "cxx_3 = ",cxx_3 , "cxxx_3 = ",cxxx_3 , "cy_3 = ",cy_3 , "cyy_3 = ",cyy_3 , "cyyy_3 = ",cyyy_3 , "cq_3 = ",cq_3 )
    print ("-------------------------------------")
    print("cc_2 = ",cc_2 , "cx_2 = ",cx_2 , "cxx_2 = ",cxx_3  , "cy_2 = ",cy_2 , "cyy_2 = ",cyy_2  , "cq_2 = ",cq_2 )

    return cc_3 , cx_3 , cxx_3 , cxxx_3, cy_3 , cyy_3 , cyyy_3 , cq_3 , cc_2 , cx_2 , cxx_2 , cy_2 , cyy_2 , cq_2
#-----------------------------------------------    
def Error_computing_cubic_equation(cc_3 , cx_3 , cxx_3 , cxxx_3, cy_3 , cyy_3 , cyyy_3 , cq_3, wb) :
    import os #for giving adress to where we want to save output files
    ws = wb.active
    name=[]
    number_of_posts = []
    number_of_followers = []
    number_of_followings = []
    number_of_likes_for_10th_ex_post = []
    number_of_likes_for_11th_ex_post = []
    number_of_likes_for_12th_ex_post = []
    number_of_self_picture_posts_form_9_previous_posts = []
    sex = []
    
    #getting data of each column of excel file into a list named by the name of corrospondant name of that column 
    for row in ws :
        name.append(row[0].value)
        number_of_posts.append(row[1].value)
        number_of_followers.append(row[2].value)
        number_of_followings.append(row[3].value)
        number_of_likes_for_10th_ex_post.append(row[4].value)
        number_of_likes_for_11th_ex_post.append(row[5].value)
        number_of_likes_for_12th_ex_post.append(row[6].value)
        number_of_self_picture_posts_form_9_previous_posts.append(row[7].value)
        sex.append(row[8].value)

    del name[0]
    del number_of_posts [0]
    del number_of_followers [0]
    del number_of_followings [0]
    del number_of_likes_for_10th_ex_post [0]
    del number_of_likes_for_11th_ex_post [0]
    del number_of_likes_for_12th_ex_post [0]
    del number_of_self_picture_posts_form_9_previous_posts [0]
    del sex [0]

    #----
    i=0
    sex_values = []
    while i < 100 :
        if sex[i] == "m" :
            sex_values.append(+10)
        elif sex[i] == "f" :
            sex_values.append(-10)
        i+=1
    #----
    i=0
    ratio_of_followings_to_followers_for_all_users = []
    while i < 100 : 
        t = number_of_followings[i]/number_of_followers[i]
        ratio_of_followings_to_followers_for_all_users.append(t)   
        i+=1
    #----
    i=0
    mean_like_for_all_users = []
    while i < 100 :
        total_like = number_of_likes_for_10th_ex_post[i] + number_of_likes_for_11th_ex_post[i] + number_of_likes_for_12th_ex_post[i]
        mean_like = total_like/3
        mean_like_for_all_users.append (mean_like)
        i+=1
    #----
    i=0
    mean_like_to_follower_for_all_users = []
    while i < 100 :
        mean_like_to_follower = mean_like_for_all_users[i]/number_of_followers[i]
        mean_like_to_follower_for_all_users.append (mean_like_to_follower)
        i+=1

    #----------------------------
    x = mean_like_to_follower_for_all_users
    y = ratio_of_followings_to_followers_for_all_users
    z = sex_values

    
    xx = [] #we want to get xx = x**2
    xxx = [] #we want to get xxx = x**3
    for x_element in x :
        element_power_2 = x_element**2
        xx.append(element_power_2)
        element_power_3 = x_element**3
        xxx.append(element_power_3)
        
    yy = [] #we want to get yy = y**2
    yyy = [] #we want to get yyy = y**3
    for y_element in y :
        element_power_2 = y_element**2
        yy.append(element_power_2)
        element_power_3 = y_element**3
        yyy.append(element_power_3)
        
    i=0
    narcissism_total = []
    while i < 100 :
        narcissism = cc_3 + cx_3*x[i] + cxx_3*xx[i] + cxxx_3*xxx[i] + cy_3*y[i] + cyy_3*yy[i] + cyyy_3*yyy[i] + cq_3*z[i]
        narcissism_total.append(narcissism)
        i+=1
    #-----------------------------
    #finding error : notice we set two group : high narcissism [5,9] (ie narcissism < %50) , low narcissism [0,4] (ie narcissism < %50), and by the above equation we find if the user belong to the first group or second group , in the following we compute our error in test dataset which is 5 in 100 means our model (above equation with the amounts we get for its variables) in 95% of cases guess true
    i=0
##    error_total = []
    error = 0
    while i < 100 :
        if narcissism_total[i] >= 5 :
            if number_of_self_picture_posts_form_9_previous_posts [i] > 4.5 :
                error += 0
            elif number_of_self_picture_posts_form_9_previous_posts [i] < 4.5 :
                error += 1
        
        elif narcissism_total[i] <= 4 :
            if number_of_self_picture_posts_form_9_previous_posts [i] < 4.5 :
                error += 0
            elif number_of_self_picture_posts_form_9_previous_posts [i] > 4.5 :
                error += 1
##        error = narcissism_total[i] - number_of_self_picture_posts_form_9_previous_posts [i]
##        error_total.append(error)
        i+=1
##    error_sum = 0
##    for e in error_total :
##        error_sum += abs(e)
##    mean_error = error_sum/100
    print("error_3 :", error)


    i=0
    error_total = []
    while i  < 100 :
        error_exact = narcissism_total[i] - number_of_self_picture_posts_form_9_previous_posts [i]
        error_total.append(error_exact)
        i+=1
    error_sum = 0
    for e in error_total :
        error_sum += abs(e)
    mean_error_exact = error_sum/100
    print("mean_error_exact_3 :",mean_error_exact)    
#-----------------------------------------------    
def Error_computing_quadratic_equation(cc_2 , cx_2 , cxx_2 , cy_2 , cyy_2 , cq_2, wb) :
    import os #for giving adress to where we want to save output files
    ws = wb.active
    name=[]
    number_of_posts = []
    number_of_followers = []
    number_of_followings = []
    number_of_likes_for_10th_ex_post = []
    number_of_likes_for_11th_ex_post = []
    number_of_likes_for_12th_ex_post = []
    number_of_self_picture_posts_form_9_previous_posts = []
    sex = []
    
    #getting data of each column of excel file into a list named by the name of corrospondant name of that column 
    for row in ws :
        name.append(row[0].value)
        number_of_posts.append(row[1].value)
        number_of_followers.append(row[2].value)
        number_of_followings.append(row[3].value)
        number_of_likes_for_10th_ex_post.append(row[4].value)
        number_of_likes_for_11th_ex_post.append(row[5].value)
        number_of_likes_for_12th_ex_post.append(row[6].value)
        number_of_self_picture_posts_form_9_previous_posts.append(row[7].value)
        sex.append(row[8].value)

    del name[0]
    del number_of_posts [0]
    del number_of_followers [0]
    del number_of_followings [0]
    del number_of_likes_for_10th_ex_post [0]
    del number_of_likes_for_11th_ex_post [0]
    del number_of_likes_for_12th_ex_post [0]
    del number_of_self_picture_posts_form_9_previous_posts [0]
    del sex [0]

    #----
    i=0
    sex_values = []
    while i < 100 :
        if sex[i] == "m" :
            sex_values.append(+10)
        elif sex[i] == "f" :
            sex_values.append(-10)
        i+=1
    #----
    i=0
    ratio_of_followings_to_followers_for_all_users = []
    while i < 100 : 
        t = number_of_followings[i]/number_of_followers[i]
        ratio_of_followings_to_followers_for_all_users.append(t)   
        i+=1
    #----
    i=0
    mean_like_for_all_users = []
    while i < 100 :
        total_like = number_of_likes_for_10th_ex_post[i] + number_of_likes_for_11th_ex_post[i] + number_of_likes_for_12th_ex_post[i]
        mean_like = total_like/3
        mean_like_for_all_users.append (mean_like)
        i+=1
    #----
    i=0
    mean_like_to_follower_for_all_users = []
    while i < 100 :
        mean_like_to_follower = mean_like_for_all_users[i]/number_of_followers[i]
        mean_like_to_follower_for_all_users.append (mean_like_to_follower)
        i+=1

    #----------------------------
    x = mean_like_to_follower_for_all_users
    y = ratio_of_followings_to_followers_for_all_users
    z = sex_values

    
    xx = [] #we want to get xx = x**2
    for x_element in x :
        element_power_2 = x_element**2
        xx.append(element_power_2)

        
    yy = [] #we want to get yy = y**2
    for y_element in y :
        element_power_2 = y_element**2
        yy.append(element_power_2)

        
    i=0
    narcissism_total = []
    while i < 100 :
        narcissism = cc_2 + cx_2*x[i] + cxx_2*xx[i] +  cy_2*y[i] + cyy_2*yy[i]  + cq_2*z[i]
        narcissism_total.append(narcissism)
        i+=1
    #-----------------------------
    #finding error : notice we set two group : high narcissism [5,9] (ie narcissism < %50) , low narcissism [0,4] (ie narcissism < %50), and by the above equation we find if the user belong to the first group or second group , in the following we compute our error in test dataset which is 5 in 100 means our model (above equation with the amounts we get for its variables) in 95% of cases guess true
    i=0
##    error_total = []
    error = 0
    while i < 100 :
        if narcissism_total[i] >= 5 :
            if number_of_self_picture_posts_form_9_previous_posts [i] > 4.5 :
                error += 0
            elif number_of_self_picture_posts_form_9_previous_posts [i] < 4.5 :
                error += 1
        
        elif narcissism_total[i] <= 4 :
            if number_of_self_picture_posts_form_9_previous_posts [i] < 4.5 :
                error += 0
            elif number_of_self_picture_posts_form_9_previous_posts [i] > 4.5 :
                error += 1
##        error = narcissism_total[i] - number_of_self_picture_posts_form_9_previous_posts [i]
##        error_total.append(error)
        i+=1
##    error_sum = 0
##    for e in error_total :
##        error_sum += abs(e)
##    mean_error = error_sum/100
    print("error_2 :", error)


    i=0
    error_total = []
    while i  < 100 :
        error_exact = narcissism_total[i] - number_of_self_picture_posts_form_9_previous_posts [i]
        error_total.append(error_exact)
        i+=1
    error_sum = 0
    for e in error_total :
        error_sum += abs(e)
    mean_error_exact = error_sum/100
    print("mean_error_exact_2 :",mean_error_exact)

#---------------------------------------
def main():
    from openpyxl import load_workbook
    wb = load_workbook('TestDataset.xlsx')
    cc_3 , cx_3 , cxx_3 , cxxx_3, cy_3 , cyy_3 , cyyy_3 , cq_3 , cc_2 , cx_2 , cxx_2 , cy_2 , cyy_2 , cq_2 = finding_coeficients()
    Error_computing_cubic_equation(cc_3 , cx_3 , cxx_3 , cxxx_3, cy_3 , cyy_3 , cyyy_3 , cq_3,wb)
    Error_computing_quadratic_equation(cc_2 , cx_2 , cxx_2 , cy_2 , cyy_2  , cq_2, wb)
#---------------------------------------
main()
input("\n press enter key to exit.")
